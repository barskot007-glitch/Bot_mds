from __future__ import annotations

import csv
from collections.abc import AsyncIterator, Iterable
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.broadcast_support import (
    BroadcastRecipient,
    SupportMessage,
    SupportTicket,
)
from app.models.enums import RegistrationStatus
from app.models.users_events import Event, Registration, User
from app.repositories.users import UserRepository
from app.utils.time import format_datetime, utcnow


class ExportService:
    """Формирует CSV/XLSX без постоянного хранения файлов на диске."""

    PAGE_SIZE = 1_000

    def __init__(
        self,
        session: AsyncSession,
        *,
        active_user_days: int = 30,
        new_user_days: int = 7,
    ) -> None:
        self.session = session
        self.active_user_days = active_user_days
        self.new_user_days = new_user_days

    @staticmethod
    def user_headers() -> list[str]:
        return [
            "Telegram ID",
            "Username",
            "Имя",
            "Фамилия",
            "Страна",
            "Телефон",
            "Email",
            "Возраст",
            "Возрастная группа",
            "История участия в проектах МДС",
            "Дата регистрации UTC",
            "Последняя активность UTC",
            "Подписан",
            "Заблокировал бота",
            "Регистраций",
            "Посещено мероприятий",
        ]

    async def iter_user_rows(
        self, filters: dict[str, Any] | None = None
    ) -> AsyncIterator[list[Any]]:
        repo = UserRepository(self.session)
        audience_filter = filters or {"subscribed_only": False}
        offset = 0
        while True:
            users = await repo.audience_page(
                audience_filter,
                now=utcnow(),
                active_days=self.active_user_days,
                new_days=self.new_user_days,
                offset=offset,
                limit=self.PAGE_SIZE,
            )
            if not users:
                return
            user_ids = [user.id for user in users]
            counts = await self.session.execute(
                select(
                    Registration.user_id,
                    func.count(Registration.id),
                    func.count(Registration.id).filter(
                        Registration.status == RegistrationStatus.ATTENDED
                    ),
                )
                .where(Registration.user_id.in_(user_ids))
                .group_by(Registration.user_id)
            )
            count_map = {
                user_id: (int(total), int(attended)) for user_id, total, attended in counts.all()
            }
            for user in users:
                total, attended = count_map.get(user.id, (0, 0))
                yield [
                    user.telegram_id,
                    user.username or "",
                    user.first_name or "",
                    user.last_name or "",
                    user.country or "",
                    user.phone or "",
                    user.email or "",
                    user.age or "",
                    user.age_group or "",
                    user.participation_history or "",
                    format_datetime(user.registered_at, "UTC"),
                    format_datetime(user.last_activity_at, "UTC"),
                    "да" if user.is_subscribed else "нет",
                    "да" if user.is_blocked else "нет",
                    total,
                    attended,
                ]
            offset += len(users)

    async def users_csv(self, filters: dict[str, Any] | None = None) -> bytes:
        buffer = StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerow(self.user_headers())
        async for row in self.iter_user_rows(filters):
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8-sig")

    async def users_xlsx(self, filters: dict[str, Any] | None = None) -> bytes:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Пользователи")
        sheet.append(self.user_headers())
        async for row in self.iter_user_rows(filters):
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def ticket_headers() -> list[str]:
        return [
            "Номер",
            "Telegram ID",
            "Username",
            "Тема",
            "Статус",
            "Назначенный администратор ID",
            "Создано UTC",
            "Последний ответ UTC",
            "Закрыто UTC",
            "Сообщений",
        ]

    async def iter_ticket_rows(self) -> AsyncIterator[list[Any]]:
        offset = 0
        while True:
            rows = (
                await self.session.execute(
                    select(
                        SupportTicket,
                        User,
                        func.count(SupportMessage.id),
                    )
                    .join(User, User.id == SupportTicket.user_id)
                    .outerjoin(SupportMessage, SupportMessage.ticket_id == SupportTicket.id)
                    .group_by(SupportTicket.id, User.id)
                    .order_by(SupportTicket.created_at.desc())
                    .offset(offset)
                    .limit(self.PAGE_SIZE)
                )
            ).all()
            if not rows:
                return
            for ticket, user, message_count in rows:
                yield [
                    ticket.number,
                    user.telegram_id,
                    user.username or "",
                    ticket.subject,
                    ticket.status.value,
                    ticket.assigned_admin_id or "",
                    format_datetime(ticket.created_at, "UTC"),
                    format_datetime(ticket.last_reply_at, "UTC"),
                    format_datetime(ticket.closed_at, "UTC"),
                    int(message_count),
                ]
            offset += len(rows)

    async def tickets_csv(self) -> bytes:
        return await self._async_csv(self.ticket_headers(), self.iter_ticket_rows())

    async def tickets_xlsx(self) -> bytes:
        return await self._async_xlsx("Обращения", self.ticket_headers(), self.iter_ticket_rows())

    @staticmethod
    def participation_headers() -> list[str]:
        return [
            "Telegram ID",
            "Username",
            "Мероприятие",
            "ID мероприятия",
            "Статус регистрации",
            "Дата регистрации UTC",
            "Дата отмены UTC",
            "Участие подтверждено UTC",
            "Комментарий администратора",
        ]

    async def iter_participation_rows(
        self, event_id: str | None = None
    ) -> AsyncIterator[list[Any]]:
        offset = 0
        while True:
            query = (
                select(Registration, User, Event)
                .join(User, User.id == Registration.user_id)
                .join(Event, Event.id == Registration.event_id)
                .order_by(Registration.registered_at.desc())
                .offset(offset)
                .limit(self.PAGE_SIZE)
            )
            if event_id:
                query = query.where(Registration.event_id == event_id)
            rows = (await self.session.execute(query)).all()
            if not rows:
                return
            for registration, user, event in rows:
                yield [
                    user.telegram_id,
                    user.username or "",
                    event.title,
                    event.id,
                    registration.status.value,
                    format_datetime(registration.registered_at, "UTC"),
                    format_datetime(registration.cancelled_at, "UTC"),
                    format_datetime(registration.attendance_confirmed_at, "UTC"),
                    registration.admin_comment or "",
                ]
            offset += len(rows)

    async def participation_csv(self, event_id: str | None = None) -> bytes:
        return await self._async_csv(
            self.participation_headers(), self.iter_participation_rows(event_id)
        )

    async def participation_xlsx(self, event_id: str | None = None) -> bytes:
        sheet = "Участники" if event_id else "История участия"
        return await self._async_xlsx(
            sheet,
            self.participation_headers(),
            self.iter_participation_rows(event_id),
        )

    @staticmethod
    def recipient_headers() -> list[str]:
        return [
            "Telegram ID",
            "Username",
            "Страна",
            "Возрастная группа",
            "Статус отправки",
            "Попыток",
            "Telegram message ID",
            "Отправлено UTC",
            "Ошибка",
        ]

    async def iter_broadcast_recipient_rows(self, broadcast_id: str) -> AsyncIterator[list[Any]]:
        offset = 0
        while True:
            rows = (
                await self.session.execute(
                    select(BroadcastRecipient, User)
                    .join(User, User.id == BroadcastRecipient.user_id)
                    .where(BroadcastRecipient.broadcast_id == broadcast_id)
                    .order_by(BroadcastRecipient.created_at.asc())
                    .offset(offset)
                    .limit(self.PAGE_SIZE)
                )
            ).all()
            if not rows:
                return
            for recipient, user in rows:
                yield [
                    user.telegram_id,
                    user.username or "",
                    user.country or "",
                    user.age_group or "",
                    recipient.status.value,
                    recipient.attempts,
                    recipient.telegram_message_id or "",
                    format_datetime(recipient.sent_at, "UTC"),
                    recipient.error_message or "",
                ]
            offset += len(rows)

    async def broadcast_recipients_csv(self, broadcast_id: str) -> bytes:
        return await self._async_csv(
            self.recipient_headers(), self.iter_broadcast_recipient_rows(broadcast_id)
        )

    async def broadcast_recipients_xlsx(self, broadcast_id: str) -> bytes:
        return await self._async_xlsx(
            "Получатели",
            self.recipient_headers(),
            self.iter_broadcast_recipient_rows(broadcast_id),
        )

    @staticmethod
    def statistics_csv(statistics: dict[str, object]) -> bytes:
        return ExportService._csv_bytes(
            ["Показатель", "Значение"],
            [[key, value] for key, value in statistics.items()],
        )

    @staticmethod
    def statistics_xlsx(title: str, statistics: dict[str, object]) -> bytes:
        return ExportService._xlsx_bytes(
            title,
            ["Показатель", "Значение"],
            [[key, value] for key, value in statistics.items()],
        )

    @staticmethod
    async def _async_csv(headers: list[str], rows: AsyncIterator[list[Any]]) -> bytes:
        buffer = StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerow(headers)
        async for row in rows:
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8-sig")

    @staticmethod
    async def _async_xlsx(
        title: str,
        headers: list[str],
        rows: AsyncIterator[list[Any]],
    ) -> bytes:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title[:31])
        sheet.append(headers)
        async for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _csv_bytes(headers: list[str], rows: Iterable[list[object]]) -> bytes:
        buffer = StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        return buffer.getvalue().encode("utf-8-sig")

    @staticmethod
    def _xlsx_bytes(title: str, headers: list[str], rows: Iterable[list[object]]) -> bytes:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title[:31])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
